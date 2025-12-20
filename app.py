import streamlit as st
import streamlit as st
import pandas as pd
from io import BytesIO
from ibm import extract_ibm_data_from_pdf, create_styled_excel, create_styled_excel_template2, correct_descriptions, extract_last_page_text
from ibm_template2 import extract_ibm_template2_from_pdf, get_extraction_debug
from template_detector import detect_ibm_template

# ✅ Must be first Streamlit command
st.set_page_config(page_title="IBM Quotation Extractor", layout="wide")

# ---------------------------
# Tool selection UI
# ---------------------------
tool_choice = st.radio(
    "Select Tool:",
    ["IBM PDF to Excel (Existing)", "IBM Excel to Excel (New)"]
)

if tool_choice == "IBM PDF to Excel (Existing)":
    # ---------------------------
    # Static content
    # ---------------------------
    compliance_text = """<Paste compliance text here>"""
    logo_path = "image.png"

    st.title("🎯 IBM Quotation PDF to Excel Converter")
    st.markdown("Upload your IBM quotation PDF - the system will automatically detect the template type")

    # Sidebar info for supported templates
    with st.sidebar:
        st.header("📋 Supported Templates")
        st.info("""
        **Auto-Detection Available:**
        
        📦 **Template 1: Parts Information**
        - Coverage dates
        - Entitled/Bid pricing
        - Parts table structure
        
        ☁️ **Template 2: Software as a Service**
        - Subscription parts
        - Service agreements
        - Commit values
        """)

    # Create two columns for layout
    col1, col2 = st.columns([2, 1])

    with col1:
        st.subheader("📁 Upload Master Price List (Optional)")
        master_csv = st.file_uploader(
            "Upload IBM Price List CSV", 
            type=["csv"], 
            key="ibm_master_csv",
            help="Upload the master CSV file to enhance quotation processing"
        )

    with col2:
        # Show upload status
        if master_csv:
            st.success("✅ Master CSV uploaded")
        else:
            st.info("📄 No master CSV uploaded")

    # Process master CSV if uploaded
    master_data = None
    if master_csv:
        try:
            master_data = pd.read_csv(master_csv)
            st.success(f"✅ Master data loaded: **{len(master_data)}** SKUs")
            with st.expander("📊 Preview Master Data"):
                st.dataframe(master_data.head(10), use_container_width=True)
        except Exception as e:
            st.error(f"❌ Error reading master CSV: {e}")

    st.markdown("---")

    # PDF Upload Section
    st.subheader("📤 Upload IBM Quotation PDF")
    uploaded_file = st.file_uploader(
        "Upload IBM Quotation PDF (Auto-detects template)", 
        type=["pdf"],
        help="Supports both Parts Information and Software as a Service templates"
    )

    if uploaded_file:
        st.success("✅ PDF uploaded successfully!")
        
        # Create columns for template detection display
        col1, col2 = st.columns([3, 1])

elif tool_choice == "IBM Excel to Excel (New)":
    st.header("🆕 IBM Excel to Excel (v2)")
    st.info("Upload an IBM quotation Excel file. The tool will extract line items from the second sheet and generate a styled quotation Excel file, matching the original format.")

    logo_path = "image.png"
    compliance_text = ""  # Add compliance text if needed

    st.subheader("📤 Upload IBM Quotation Excel File")

    uploaded_excel = st.file_uploader(
        "Upload IBM Quotation Excel (.xlsx, .xlsm, .xls)",
        type=["xlsx", "xlsm", "xls"],
        help="Supports .xlsx, .xlsm, and .xls files. The tool will extract line items from the second sheet using the specified column mapping."
    )


    if uploaded_excel:
        import pandas as pd
        from sales.ibm_v2 import create_styled_excel_v2
        import io
        file_type = uploaded_excel.type
        file_name = uploaded_excel.name.lower()
        df = None
        # Always use pandas/xlrd for .xls
        if file_type == "application/vnd.ms-excel" or file_name.endswith(".xls"):
            try:
                xls = pd.ExcelFile(uploaded_excel, engine="xlrd")
                if len(xls.sheet_names) < 2:
                    st.error("❌ The uploaded Excel file does not have a second sheet.")
                    st.stop()
                df = xls.parse(xls.sheet_names[1], header=None)
            except Exception as e:
                st.error(f"❌ Error reading .xls file: {e}")
                st.stop()
        else:
            # Try openpyxl for .xlsx/.xlsm, fallback to pandas if it fails
            try:
                from openpyxl import load_workbook
                wb = load_workbook(uploaded_excel, data_only=True)
                sheetnames = wb.sheetnames
                if len(sheetnames) < 2:
                    st.error("❌ The uploaded Excel file does not have a second sheet.")
                    st.stop()
                ws = wb[sheetnames[1]]
                data_rows = list(ws.values)
                df = pd.DataFrame(data_rows)
            except Exception as e:
                # Fallback: try pandas
                try:
                    xls = pd.ExcelFile(uploaded_excel)
                    if len(xls.sheet_names) < 2:
                        st.error("❌ The uploaded Excel file does not have a second sheet.")
                        st.stop()
                    df = xls.parse(xls.sheet_names[1], header=None)
                except Exception as e2:
                    st.error(f"❌ Error reading Excel file: {e}\nFallback also failed: {e2}")
                    st.stop()

        def safe_cell(df, row, col):
            try:
                return df.iloc[row, col] if row < len(df.index) and col < len(df.columns) else ""
            except Exception:
                return ""

        header_info = {
            "Customer Name": safe_cell(df, 3, 2),
            "Bid Number": safe_cell(df, 4, 2),
            "PA Agreement Number": safe_cell(df, 5, 2),
            "PA Site Number": safe_cell(df, 6, 2),
            "Select Territory": safe_cell(df, 7, 2),
            "Government Entity (GOE)": safe_cell(df, 8, 2),
            "Reseller Name": safe_cell(df, 9, 2),
            "City": safe_cell(df, 10, 2),
            "Country": safe_cell(df, 11, 2),
            "Maximum End User Price (MEP)": safe_cell(df, 12, 2),
            "Bid Expiration Date": safe_cell(df, 13, 2),
        }

        data = []
        for i in range(9, len(df)):
            row = df.iloc[i]
            sku = row[9] if len(row) > 9 else ""
            desc = row[1] if len(row) > 1 else ""
            qty = row[6] if len(row) > 6 else ""
            start_date = row[7] if len(row) > 7 else ""
            end_date = row[20] if len(row) > 20 else ""
            cost = row[7] if len(row) > 7 else ""
            if sku:
                data.append([sku, desc, qty, start_date, end_date, cost])

        st.info(f"Extracted {len(data)} line items from Excel.")
        if data:
            st.dataframe(pd.DataFrame(data, columns=["SKU", "Description", "Quantity", "Start Date", "End Date", "Cost (AED)"]), use_container_width=True)

            if st.button("🎯 Generate Styled Quotation Excel", type="primary", use_container_width=True, key="generate_excel_v2_btn"):
                with st.spinner("📊 Creating styled Excel quotation..."):
                    try:
                        output = BytesIO()
                        logo_path = "image.png"
                        compliance_text = ""
                        ibm_terms_text = ""
                        create_styled_excel_v2(
                            data,
                            header_info,
                            logo_path,
                            output,
                            compliance_text,
                            ibm_terms_text
                        )
                        st.download_button(
                            label="📥 Download Excel Quotation",
                            data=output.getvalue(),
                            file_name="IBM_Quotation_v2.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True
                        )
                        st.success("✅ Excel file generated successfully!")
                        st.balloons()
                    except Exception as e:
                        st.error(f"❌ Error generating Excel: {str(e)}")
                        st.exception(e)


    # (Removed duplicate/old upload logic for Excel files. Only robust, type-checked logic remains above.)

            st.info(f"Extracted {len(data)} line items from Excel.")
            if data:
                df = pd.DataFrame(data, columns=["SKU", "Description", "Quantity", "Start Date", "End Date", "Cost (AED)"])
                st.dataframe(df, use_container_width=True)

                # Generate Excel
                if st.button("🎯 Generate Styled Quotation Excel", type="primary", use_container_width=True):
                    with st.spinner("📊 Creating styled Excel quotation..."):
                        try:
                            output = BytesIO()
                            # For v2, cost is in AED, need to convert to USD for the output (handled in v2 function)
                            from terms_template import get_terms_section
                            ibm_terms_text = ""  # Add logic to extract if needed
                            create_styled_excel_v2(
                                data,
                                header_info,
                                logo_path,
                                output,
                                compliance_text,
                                ibm_terms_text
                            )
                            st.download_button(
                                label="📥 Download Excel Quotation",
                                data=output.getvalue(),
                                file_name="IBM_Quotation_v2.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                use_container_width=True
                            )
                            st.success("✅ Excel file generated successfully!")
                            st.balloons()
                        except Exception as e:
                            st.error(f"❌ Error generating Excel: {str(e)}")
                            st.exception(e)


# ✅ Must be first Streamlit command
# st.set_page_config(page_title="IBM Quotation Extractor", layout="wide")

# # ---------------------------
# # Static content
# # ---------------------------
# compliance_text = """<Paste compliance text here>"""
# logo_path = "image.png"

# st.title("🎯 IBM Quotation PDF to Excel Converter")
# st.markdown("Upload your IBM quotation PDF - the system will automatically detect the template type")

# # Sidebar info for supported templates
# with st.sidebar:
#     st.header("📋 Supported Templates")
#     st.info("""
#     **Auto-Detection Available:**
    
#     📦 **Template 1: Parts Information**
#     - Coverage dates
#     - Entitled/Bid pricing
#     - Parts table structure
    
#     ☁️ **Template 2: Software as a Service**
#     - Subscription parts
#     - Service agreements
#     - Commit values
#     """)

# # Create two columns for layout
# col1, col2 = st.columns([2, 1])

# with col1:
#     st.subheader("📁 Upload Master Price List (Optional)")
#     master_csv = st.file_uploader(
#         "Upload IBM Price List CSV", 
#         type=["csv"], 
#         key="ibm_master_csv",
#         help="Upload the master CSV file to enhance quotation processing"
#     )

# with col2:
#     # Show upload status
#     if master_csv:
#         st.success("✅ Master CSV uploaded")
#     else:
#         st.info("📄 No master CSV uploaded")

# # Process master CSV if uploaded
# master_data = None
# if master_csv:
#     try:
#         master_data = pd.read_csv(master_csv)
#         st.success(f"✅ Master data loaded: **{len(master_data)}** SKUs")
#         with st.expander("📊 Preview Master Data"):
#             st.dataframe(master_data.head(10), use_container_width=True)
#     except Exception as e:
#         st.error(f"❌ Error reading master CSV: {e}")

# st.markdown("---")

# # PDF Upload Section
# st.subheader("📤 Upload IBM Quotation PDF")
# uploaded_file = st.file_uploader(
#     "Upload IBM Quotation PDF (Auto-detects template)", 
#     type=["pdf"],
#     help="Supports both Parts Information and Software as a Service templates"
# )

# if uploaded_file:
#     st.success("✅ PDF uploaded successfully!")
    
#     # Create columns for template detection display
#     col1, col2 = st.columns([3, 1])
    
#     with col1:
#         # Detect template type
#         file_bytes = uploaded_file.read()
#         uploaded_file.seek(0)  # Reset file pointer
        
#         template_type = detect_ibm_template(BytesIO(file_bytes))
        
#         # Set template info based on detection
#         if template_type == 'template2':
#             template_info = {
#                 'name': 'IBM Software as a Service',
#                 'description': 'Subscription-based service quotation',
#                 'icon': '☁️'
#             }
#         else:
#             template_info = {
#                 'name': 'IBM Parts Information',
#                 'description': 'Parts-based quotation with coverage dates',
#                 'icon': '📦'
#             }
        
#         # Show detected template with nice styling
#         st.markdown(f"""
#         <div style="
#             padding: 1rem; 
#             border-radius: 0.5rem; 
#             background: linear-gradient(90deg, #e8f4fd, #f0f9ff);
#             border-left: 4px solid #1f77b4;
#             margin: 1rem 0;
#         ">
#             <h3 style="margin: 0; color: #1f77b4;">
#                 {template_info['icon']} {template_info['name']}
#             </h3>
#             <p style="margin: 0.5rem 0 0 0; color: #666;">
#                 📋 {template_info['description']}
#             </p>
#         </div>
#         """, unsafe_allow_html=True)
    
#     with col2:
#         st.metric("Detected Template", template_type.upper())
    
#     st.markdown("---")
    
#     # Extract data from PDF
#     with st.spinner("📄 Extracting data from PDF..."):
#         try:
#             # Use appropriate extractor based on template
#             if template_type == 'template2':
#                 st.write(f"🔍 DEBUG: Using Template 2 extractor")
#                 data, header_info = extract_ibm_template2_from_pdf(BytesIO(file_bytes))
#                 debug_messages = get_extraction_debug()
#                 st.write(f"🔍 DEBUG: Got {len(debug_messages)} debug messages")
#                 st.write(f"🔍 DEBUG: Extracted {len(data)} line items")
#                 ibm_terms_text = extract_last_page_text(BytesIO(file_bytes))
#             else:
#                 data, header_info = extract_ibm_data_from_pdf(BytesIO(file_bytes))
#                 ibm_terms_text = extract_last_page_text(BytesIO(file_bytes))
#                 debug_messages = ["Template 1 extraction - check pdf_extraction_debug.log for details"]
            
#         except Exception as e:
#             st.error(f"❌ Error extracting data: {str(e)}")
#             import traceback
#             st.error(f"Full traceback: {traceback.format_exc()}")
#             data = []
#             header_info = {}
#             debug_messages = [f"Error: {str(e)}", traceback.format_exc()]
    
#     if data and len(data) > 0:
#         # Success metrics
#         col1, col2, col3 = st.columns(3)
#         with col1:
#             st.metric("📄 Template", template_type.title())
#         with col2:
#             st.metric("📦 Line Items", len(data))
#         with col3:
#             total_value = sum([row[6] for row in data if len(row) > 6 and row[6]])
#             st.metric("💰 Total Value", f"AED {total_value:,.2f}")
        
#         st.success(f"✅ Successfully extracted **{len(data)}** line items from {template_info['name']}")
        
#         # Show header information
#         with st.expander("📋 Quotation Information", expanded=True):
#             col1, col2, col3 = st.columns(3)
            
#             with col1:
#                 st.markdown("**Customer Details:**")
#                 st.text(f"Name: {header_info.get('Customer Name', 'N/A')}")
#                 st.text(f"Reseller: {header_info.get('Reseller Name', 'N/A')}")
                
#             with col2:
#                 st.markdown("**Bid Information:**")
#                 st.text(f"Bid Number: {header_info.get('Bid Number', 'N/A')}")
#                 st.text(f"PA Agreement: {header_info.get('PA Agreement Number', 'N/A')}")
                
#             with col3:
#                 st.markdown("**IBM Details:**")
#                 st.text(f"Opportunity: {header_info.get('IBM Opportunity Number', 'N/A')}")
#                 st.text(f"Territory: {header_info.get('Select Territory', 'N/A')}")

#         # 🔍 Debug Log Viewer for Template 2
#         if template_type == 'template2':
#             with st.expander("🔍 Detailed Extraction Log (Template 2)", expanded=False):
#                 st.markdown("### 📝 Debug Information")
#                 st.info("This log shows exactly what the extractor found in the PDF and how it processed each line item.")
                
#                 # Show log in a text area
#                 if debug_messages:
#                     log_text = "\n".join(debug_messages)
#                     st.text_area(
#                         "Extraction Debug Log",
#                         value=log_text,
#                         height=600,
#                         help="Detailed step-by-step extraction log"
#                     )
                    
#                     # Download button for log
#                     st.download_button(
#                         label="📥 Download Full Debug Log",
#                         data=log_text,
#                         file_name="template2_extraction_log.txt",
#                         mime="text/plain",
#                         help="Save the complete extraction log for analysis"
#                     )
                    
#                     # Show log file path
#                     st.info("💾 A copy is also saved to: `template2_extraction_debug.log`")
#                 else:
#                     st.warning("No debug information available")
        
#         # 🔍 Raw Extraction Analysis
#         with st.expander("🔍 Raw Extraction Analysis", expanded=False):
#             st.markdown("### 📊 Extraction Details")
            
#             # Show extraction statistics
#             col1, col2, col3, col4 = st.columns(4)
#             with col1:
#                 st.metric("Total Rows", len(data))
#             with col2:
#                 valid_prices = [row[6] for row in data if len(row) > 6 and row[6] is not None]
#                 st.metric("Valid Prices", len(valid_prices))
#             with col3:
#                 total_qty = sum([row[2] for row in data if len(row) > 2 and row[2] is not None])
#                 st.metric("Total Quantity", total_qty)
#             with col4:
#                 unique_skus = set([row[0] for row in data if len(row) > 0 and row[0]])
#                 st.metric("Unique SKUs", len(unique_skus))
            
#             # Raw data table
#             st.markdown("### 📋 Raw Extracted Data")
#             if data:
#                 # Create DataFrame based on template type
#                 if template_type == 'template2':
#                     # Template 2: [sku, desc, qty, duration, start_date, end_date, bid_unit_aed, bid_total_aed, partner_price_aed]
#                     df_raw = pd.DataFrame(data, columns=[
#                         "SKU", "Description", "Quantity", "Duration", "Start Date", "End Date",
#                         "Unit Price (AED)", "Total Price (AED)", "Partner Price (AED)"
#                     ])
#                 else:
#                     # Template 1: [sku, desc, qty, start_date, end_date, bid_unit_aed, bid_total_aed]
#                     df_raw = pd.DataFrame(data, columns=[
#                         "SKU", "Description", "Quantity", "Start Date", "End Date",
#                         "Unit Price (AED)", "Total Price (AED)"
#                     ])
                
#                 # Add row numbers for reference
#                 df_raw.index = range(1, len(df_raw) + 1)
#                 st.dataframe(df_raw, use_container_width=True, height=300)
                
#                 # Export raw data option
#                 csv_raw = df_raw.to_csv(index=True)
#                 st.download_button(
#                     label="📥 Download Raw Data (CSV)",
#                     data=csv_raw,
#                     file_name="ibm_raw_extraction.csv",
#                     mime="text/csv",
#                     help="Download the raw extracted data for analysis"
#                 )
            
#             # Header information details
#             st.markdown("### 📝 Extracted Header Information")
#             if header_info:
#                 header_df = pd.DataFrame([
#                     {"Field": key, "Value": value} 
#                     for key, value in header_info.items()
#                 ])
#                 st.dataframe(header_df, use_container_width=True)
                
#                 # Export header info
#                 csv_header = header_df.to_csv(index=False)
#                 st.download_button(
#                     label="📥 Download Header Info (CSV)",
#                     data=csv_header,
#                     file_name="ibm_header_info.csv",
#                     mime="text/csv",
#                     help="Download the extracted header information"
#                 )
            
#             # IBM Terms preview
#             st.markdown("### 📄 IBM Terms Preview")
#             if 'ibm_terms_text' in locals() and ibm_terms_text:
#                 terms_lines = ibm_terms_text.split('\n')
#                 st.text_area(
#                     "IBM Terms Content (First 500 chars):",
#                     value=ibm_terms_text[:500] + "..." if len(ibm_terms_text) > 500 else ibm_terms_text,
#                     height=100,
#                     disabled=True
#                 )
#                 st.info(f"📊 Total IBM Terms length: {len(ibm_terms_text)} characters, {len(terms_lines)} lines")
#             else:
#                 st.warning("⚠️ No IBM Terms extracted from last page")

#         # Continue with existing code...
#                 st.text(f"City: {header_info.get('City', 'N/A')}")
#                 st.text(f"Country: {header_info.get('Country', 'N/A')}")
            
#             with col2:
#                 st.markdown("**Bid Information:**")
#                 st.text(f"Bid Number: {header_info.get('Bid Number', 'N/A')}")
#                 st.text(f"PA Agreement: {header_info.get('PA Agreement Number', 'N/A')}")
#                 st.text(f"PA Site: {header_info.get('PA Site Number', 'N/A')}")
            
#             with col3:
#                 st.markdown("**Other Details:**")
#                 st.text(f"Reseller: {header_info.get('Reseller Name', 'N/A')}")
#                 st.text(f"Territory: {header_info.get('Select Territory', 'N/A')}")
#                 st.text(f"GOE: {header_info.get('Government Entity (GOE)', 'N/A')}")
        
#         # Show extraction debug info
#         with st.expander("🔍 Debug: PDF Extraction Results"):
#             st.write(f"**Template detected:** {template_info['name']}")
#             st.write(f"**Total rows extracted from PDF:** {len(data)}")
#             if data:
#                 st.write("**Extracted SKUs and descriptions:**")
#                 for i, row in enumerate(data):
#                     desc_preview = row[1][:50] + '...' if len(row[1]) > 50 else row[1]
#                     st.write(f"Row {i+1}: `{row[0]}` - {desc_preview}")
#             else:
#                 st.error("❌ No data extracted from PDF!")
        
#         # Show extracted data preview
#         with st.expander("📊 Preview Extracted Line Items"):
#             # Use different column headers based on template type
#             if template_type == 'template2':
#                 # Template 2: Show without Start/End Date
#                 preview_data = []
#                 for row in data:
#                     preview_row = [row[0], row[1], row[2], row[3], row[6], row[7]]  # Skip start_date, end_date
#                     preview_data.append(preview_row)
                
#                 preview_columns = [
#                     "SKU", "Description", "Qty", "Duration", "Unit Price AED", "Total Price AED"
#                 ]
#                 preview_df = pd.DataFrame(preview_data, columns=preview_columns)
#             else:
#                 preview_columns = [
#                     "SKU", "Description", "Qty", "Start Date", "End Date", "Unit Price AED", "Total Price AED"
#                 ]
#                 preview_df = pd.DataFrame(data, columns=preview_columns)
            
#             st.dataframe(preview_df, use_container_width=True)
        
#         # Master CSV analysis
#         if master_data is not None:
#             with st.expander("🔍 Debug: Master CSV Analysis"):
#                 st.write(f"**Total master records:** {len(master_data)}")
                
#                 # Show matches between PDF and Master
#                 pdf_skus = [row[0] for row in data]
#                 matched_skus = [sku for sku in pdf_skus if sku in master_data['SKU'].values]
#                 unmatched_skus = [sku for sku in pdf_skus if sku not in master_data['SKU'].values]
                
#                 col1, col2 = st.columns(2)
#                 with col1:
#                     st.metric("✅ Matched SKUs", len(matched_skus))
#                     if matched_skus:
#                         st.write("**Found in master:**")
#                         for sku in matched_skus[:10]:  # Show first 10
#                             st.write(f"• {sku}")
#                         if len(matched_skus) > 10:
#                             st.write(f"... and {len(matched_skus) - 10} more")
                
#                 with col2:
#                     st.metric("❌ Unmatched SKUs", len(unmatched_skus))
#                     if unmatched_skus:
#                         st.warning("**These SKUs will have blank descriptions:**")
#                         for sku in unmatched_skus[:10]:  # Show first 10
#                             st.write(f"• {sku}")
#                         if len(unmatched_skus) > 10:
#                             st.write(f"... and {len(unmatched_skus) - 10} more")
        
#         st.markdown("---")
        
#         # Apply description corrections (only for Template 1)
#         if template_type == 'template1':
#             with st.spinner("🔄 Applying description corrections..."):
#                 corrected_data = correct_descriptions(data, master_data=master_data)
#         else:
#             # For Template 2, use data as-is (descriptions are already complete)
#             corrected_data = data
#             st.info("📝 Template 2: Using extracted descriptions as-is (no CSV correction needed)")
        
#         # 📊 Data Correction Analysis (only for Template 1)
#         if template_type == 'template1':
#             with st.expander("🔄 Description Correction Analysis", expanded=False):
#                 st.markdown("### 📈 Correction Summary")
                
#                 # Calculate correction statistics
#                 corrections_made = 0
#                 corrections_blank = 0
#                 no_corrections = 0
                
#                 correction_details = []
                
#                 for i, (original_row, final_row) in enumerate(zip(data, corrected_data)):
#                     original_desc = original_row[1] if len(original_row) > 1 else ""
#                     final_desc = final_row[1] if len(final_row) > 1 else ""
#                     sku = final_row[0] if len(final_row) > 0 else ""
                    
#                     if final_desc and final_desc != original_desc:
#                         corrections_made += 1
#                         status = "✅ Updated from Master CSV"
#                         color = "green"
#                     elif not final_desc:
#                         corrections_blank += 1
#                         status = "⚠️ Set to blank (SKU not found)"
#                         color = "orange"
#                     else:
#                         no_corrections += 1
#                         status = "📄 No change needed"
#                         color = "blue"
                    
#                     correction_details.append({
#                         "Row": i + 1,
#                         "SKU": sku,
#                         "Status": status,
#                         "Original Description": original_desc[:50] + "..." if len(original_desc) > 50 else original_desc,
#                         "Final Description": final_desc[:50] + "..." if len(final_desc) > 50 else final_desc,
#                         "Color": color
#                     })
                
#                 # Display summary metrics
#                 col1, col2, col3, col4 = st.columns(4)
#                 with col1:
#                     st.metric("✅ Corrected", corrections_made)
#                 with col2:
#                     st.metric("⚠️ Set to Blank", corrections_blank)
#                 with col3:
#                     st.metric("📄 No Change", no_corrections)
#                 with col4:
#                     st.metric("📊 Total Rows", len(data))
                
#                 # Show detailed correction table
#                 st.markdown("### 📋 Detailed Correction Log")
#                 if correction_details:
#                     correction_df = pd.DataFrame(correction_details)
                    
#                     # Color-code the status
#                     def highlight_status(row):
#                         color_map = {
#                             'green': 'background-color: #d4edda; color: #155724;',
#                             'orange': 'background-color: #fff3cd; color: #856404;',
#                             'blue': 'background-color: #d1ecf1; color: #0c5460;'
#                         }
#                         return [''] * len(row) if row['Color'] not in color_map else [color_map[row['Color']] if i == 2 else '' for i in range(len(row))]
                    
#                     # Display without the color column
#                     display_df = correction_df.drop('Color', axis=1)
#                     st.dataframe(display_df, use_container_width=True, height=400)
                    
#                     # Export correction log
#                     csv_corrections = correction_df.drop('Color', axis=1).to_csv(index=False)
#                     st.download_button(
#                         label="📥 Download Correction Log (CSV)",
#                         data=csv_corrections,
#                         file_name="ibm_correction_log.csv",
#                         mime="text/csv",
#                         help="Download detailed log of all description corrections"
#                     )
                
#                 # Master CSV usage statistics
#                 if master_data is not None:
#                     st.markdown("### 📊 Master CSV Usage Stats")
#                     pdf_skus = [row[0] for row in data if len(row) > 0]
#                     total_master_skus = len(master_data)
#                     used_skus = [sku for sku in pdf_skus if sku in master_data['SKU'].values]
                    
#                     col1, col2, col3 = st.columns(3)
#                     with col1:
#                         st.metric("📦 Total Master SKUs", total_master_skus)
#                     with col2:
#                         st.metric("🎯 SKUs Used", len(used_skus))
#                     with col3:
#                         usage_rate = (len(used_skus) / len(pdf_skus) * 100) if pdf_skus else 0
#                         st.metric("📈 Usage Rate", f"{usage_rate:.1f}%")
#         else:
#             # Template 2 - Show different analysis
#             with st.expander("📊 Template 2 Data Analysis", expanded=False):
#                 st.markdown("### 📝 Description Analysis")
#                 st.info("Template 2 uses complete service blocks as descriptions - no CSV correction needed")
                
#                 # Show description lengths
#                 desc_lengths = [len(row[1]) if len(row) > 1 and row[1] else 0 for row in data]
#                 if desc_lengths:
#                     col1, col2, col3 = st.columns(3)
#                     with col1:
#                         st.metric("📏 Avg Description Length", f"{sum(desc_lengths) / len(desc_lengths):.0f} chars")
#                     with col2:
#                         st.metric("📄 Max Description Length", f"{max(desc_lengths)} chars")
#                     with col3:
#                         st.metric("📋 Min Description Length", f"{min(desc_lengths)} chars")
        
#         # Show corrected data
#         st.subheader("📊 Final BoQ Data")
        
#         # Use different column headers based on template type
#         if template_type == 'template2':
#             # Template 2: [sku, desc, qty, duration, start_date, end_date, bid_unit_aed, bid_total_aed, partner_price_aed]
#             # Display: SKU, Description, Quantity, Duration, Unit Price, Total Price, Partner Price
#             display_data = []
#             for row in corrected_data:
#                 # Include Partner Price (index 8) in display
#                 display_row = [row[0], row[1], row[2], row[3], row[6], row[7], row[8] if len(row) > 8 else None]
#                 display_data.append(display_row)
            
#             columns = [
#                 "SKU", "Product Description", "Quantity", "Duration", 
#                 "Unit Price in AED", "Total Price in AED", "Partner Price in AED"
#             ]
#             final_df = pd.DataFrame(display_data, columns=columns)
#         else:
#             # Template 1: [sku, desc, qty, start_date, end_date, bid_unit_aed, bid_total_aed]
#             columns = [
#                 "SKU", "Product Description", "Quantity", "Start Date", "End Date",
#                 "Unit Price in AED", "Total Price in AED"
#             ]
#             final_df = pd.DataFrame(corrected_data, columns=columns)
#         st.dataframe(final_df, use_container_width=True)
        
#         # Show description correction summary (Template 1 only)
#         if master_data is not None and template_type == 'template1':
#             with st.expander("🔍 Debug: Description Correction Summary"):
#                 for i, (original_row, final_row) in enumerate(zip(data, corrected_data)):
#                     original_desc = original_row[1]
#                     final_desc = final_row[1]
                    
#                     if final_desc and final_desc != original_desc:
#                         st.success(f"Row {i+1} - SKU `{final_row[0]}`: ✅ Updated from master CSV")
#                     elif not final_desc:
#                         st.warning(f"Row {i+1} - SKU `{final_row[0]}`: ⚠️ Set to blank")
#                     else:
#                         st.info(f"Row {i+1} - SKU `{final_row[0]}`: 📄 No change")
        
#         st.markdown("---")
        
#         # Generate Excel Section
#         st.subheader("📥 Generate Excel Quotation")
        
#         col1, col2 = st.columns([3, 1])
        
#         with col1:
#             if st.button("🎯 Generate Excel File", type="primary", use_container_width=True):
#                 with st.spinner("📊 Creating styled Excel quotation..."):
#                     try:
#                         output = BytesIO()
                        
#                         if template_type == 'template2':
#                             # Use dedicated Template 2 Excel generator with full 9-column data
#                             # Template 2 data: [sku, desc, qty, duration, start_date, end_date, bid_unit_aed, bid_total_aed, partner_price_aed]
#                             st.info("🎯 Using Template 2 Excel generation function...")
#                             create_styled_excel_template2(
#                                 corrected_data, 
#                                 header_info, 
#                                 logo_path, 
#                                 output, 
#                                 compliance_text, 
#                                 ibm_terms_text
#                             )
#                         else:
#                             # Use Template 1 Excel generator
#                             create_styled_excel(
#                                 corrected_data, 
#                                 header_info, 
#                                 logo_path, 
#                                 output, 
#                                 compliance_text, 
#                                 ibm_terms_text
#                             )                        # Download button
#                         bid_number = header_info.get('Bid Number', 'output')
#                         filename = f"IBM_Quotation_{bid_number}.xlsx"
                        
#                         st.download_button(
#                             label="📥 Download Excel Quotation",
#                             data=output.getvalue(),
#                             file_name=filename,
#                             mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
#                             use_container_width=True
#                         )
                        
#                         st.success("✅ Excel file generated successfully!")
#                         st.balloons()
                        
#                     except Exception as e:
#                         st.error(f"❌ Error generating Excel: {str(e)}")
#                         st.exception(e)
        
#         with col2:
#             st.metric("📄 Items to Export", len(corrected_data))
        
#         # Debug information (optional)
#         with st.expander("🔍 Advanced Debug Information", expanded=False):
#             st.markdown("**Template Detection Log:**")
#             st.write(f"Detected: {template_type} - {template_info['name']}")
            
#             st.markdown("**Extraction Log:**")
#             for msg in debug_messages[-20:]:  # Show last 20 debug messages
#                 st.text(msg)
            
#             # ADD THIS: Total Price AED Debugging
#             st.markdown("**💰 Total Price AED Calculation Debug:**")
#             if corrected_data:
#                 for i, row in enumerate(corrected_data[:5]):  # Show first 5 rows
#                     sku = row[0]
#                     qty = row[2]
#                     cost = row[6] if len(row) > 6 else 0  # bid_ext_svp_aed
                    
#                     # Calculate what should happen
#                     calculated_total = round(cost * 3.6725, 2) if cost else 0
#                     unit_price = round(calculated_total / qty, 2) if qty > 0 else 0
                    
#                     st.write(f"**Row {i+1} - SKU: {sku}**")
#                     st.write(f"  • Cost (bid_ext_svp_aed): {cost}")
#                     st.write(f"  • Quantity: {qty}")
#                     st.write(f"  • Total Price AED = {cost} × 3.6725 = {calculated_total}")
#                     st.write(f"  • Unit Price AED = {calculated_total} ÷ {qty} = {unit_price}")
#                     st.write("---")
            
#             st.markdown("**Header Info:**")
#             st.json(header_info)
    
#     else:
#         # Error handling
#         st.error("❌ No data could be extracted from the PDF")
        
#         st.markdown("### 🔧 Troubleshooting")
#         st.warning("**Please check if:**")
#         st.write("• PDF format matches supported templates")
#         st.write("• PDF contains recognizable SKUs and line items")
#         st.write("• PDF is not corrupted or password-protected")
        
#         if template_type == 'template2':
#             st.write("• PDF contains 'Software as a Service' sections")
#             st.write("• Subscription parts are clearly defined")
#         else:
#             st.write("• PDF contains 'Parts Information' table")
#             st.write("• Coverage dates are present")
        
#         with st.expander("🔍 Debug Information", expanded=True):
#             st.markdown("**Template Detection:**")
#             st.write(f"Detected: {template_type} - {template_info['name']}")
            
#             st.markdown("**Extraction Log:**")
#             for msg in debug_messages:
#                 st.text(msg)

# else:
#     # No file uploaded
#     st.info("📤 Please upload a PDF file to begin.")